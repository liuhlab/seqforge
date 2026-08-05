"""``harvest normalize`` — build the canonical text that spans are computed against.

Span verification is the hallucination tripwire, and it dies on raw PDF text: a naive grep for a
quote fails on soft hyphens, ligatures (``ﬁ`` vs ``fi``), non-breaking spaces, smart quotes, and
mid-sentence line breaks — so a *truthful* quote would be rejected and the tripwire would train us to
ignore it. The fix is to extract **once** into a normalized canonical text, store offsets
into **that**, and verify against **that**.

Deterministic and LLM-free. ``normalizer_version`` is folded into the artifact cache key, because a
normalization change silently invalidates every offset computed under the old one.
"""

from __future__ import annotations

import bisect
import hashlib
import math
import re
import unicodedata
from collections import Counter
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from ..models.records import ArchiveRecord
from .fields import DocRole, DocScope

#: CalVer YYYY.M.PATCH; bump when normalization changes (it re-defines the span space). Bumped from
#: 2026.7.0 for the invalid-unicode scrub and the page-aware PDF path — both change the canonical text
#: a quote is matched against, so every offset computed under the old value must be recomputed.
NORMALIZER_VERSION = "2026.7.1"

#: Which extractor turned a PDF into text. ``pymupdf`` (AGPL-3.0) is the DEFAULT — it read every one
#: of the ten real worm manuscripts, including two whose fonts make ``pypdf`` (BSD) raise
#: ``PdfReadError``. ``pypdf`` stays available as the permissive fallback the failure message points
#: to. Reading order is content-stream order for BOTH (no geometric sort — that interleaves columns);
#: the backend choice is really "which engine opens the file", and pymupdf opens more of them.
PdfBackend = Literal["pypdf", "pymupdf"]

#: The parse default. Chosen by an end-to-end harvest-quality eval, not by license or speed — see the
#: ``PyMuPDF`` entry in ``pyproject.toml``. Code owns this decision and it is reproducible; it is never
#: the model's to make (how to parse is byte-decided, like ``backend.params``).
DEFAULT_PDF_BACKEND: PdfBackend = "pymupdf"

# Ligatures NFKC does not decompose the way a grep needs.
_LIGATURES = {
    "ﬀ": "ff", "ﬁ": "fi", "ﬂ": "fl", "ﬃ": "ffi", "ﬄ": "ffl",
    "ﬅ": "st", "ﬆ": "st", "Ĳ": "IJ", "ĳ": "ij", "Œ": "OE", "œ": "oe",
}  # fmt: skip
# Punctuation a PDF renders "prettily" but nobody types when quoting.
_PUNCT = {
    "‘": "'", "’": "'", "‚": "'", "‛": "'",
    "“": '"', "”": '"', "„": '"', "‟": '"',
    "‐": "-", "‑": "-", "‒": "-", "–": "-", "—": "-", "―": "-",
    "′": "'", "″": '"', "­": "",  # prime, double-prime, SOFT HYPHEN (delete)
    " ": " ", " ": " ", " ": " ", " ": " ", " ": " ", "​": "",
}  # fmt: skip

# A hyphen before a line break is ambiguous, and guessing wrong corrupts the span space:
#   "chemi-\nstry"  -> the hyphen is a WRAP artifact  -> close it up  ("chemistry")
#   "3-\nprime"     -> the hyphen is SEMANTIC         -> keep it     ("3-prime")
# Digit-adjacent hyphens (3-prime, R64-1-1, 10-fold) are part of the token, so only an
# alphabetic-alphabetic break is treated as hyphenation to undo.
_WRAP_HYPHEN = re.compile(r"(?<=[A-Za-z])-\n(?=[A-Za-z])")
_KEPT_HYPHEN = re.compile(r"(?<=\w)-\n(?=\w)")
_PARA_BREAK = re.compile(r"\n[ \t]*\n+")  # blank line(s) = a real paragraph boundary
_LINE_BREAK = re.compile(r"[ \t]*\n[ \t]*")  # a lone newline inside a paragraph = a wrap artifact
_WS_RUN = re.compile(r"[ \t]{2,}")
_PARA_TOKEN = "\x00PARA\x00"

# NUL and the lone UTF-16 surrogates (U+D800–U+DFFF) are not prose: a PDF extractor emits them on a
# bad font or a figure, and they have no business in the canonical string. A lone surrogate also
# breaks the very next thing we do to it — ``"…".encode()`` for the content hash raises
# UnicodeEncodeError on one — so the scrub runs FIRST in ``normalize_text``, before any hashing or
# matching. (paper-qa added the identical scrub only after a production crash on an orphaned low
# surrogate from a single arXiv figure; better to have it before ours.)
_INVALID_UNICODE = re.compile("[\x00\ud800-\udfff]")


def clean_invalid_unicode(text: str) -> str:
    """Drop NUL bytes and orphaned UTF-16 surrogates — characters that are never legitimate prose."""
    return _INVALID_UNICODE.sub("", text)


@dataclass(frozen=True)
class PageSpan:
    """Where one physical PDF page landed in the canonical text — a half-open ``[start, end)`` range.

    It lets ``verify`` map a computed character offset back to a 1-indexed page (``SourceSpan.page``)
    without the model ever counting, and *without changing what a quote greps into*: the canonical
    text is still the ordered concatenation of the pages, so a span check is unaffected.
    """

    number: int
    start: int
    end: int


@dataclass(frozen=True)
class DeclaredSpan:
    """Where a record re-rendered one of its OWN typed columns into its own free text.

    A half-open ``[start, end)`` range into the canonical text, plus the proof that it is one: the
    attribute the submitter typed, and the value they typed into it. The proof is the whole point —
    this is not a guess about which part of a title looks like metadata, it is a byte-equal match
    against a column of the same record, so a reader can check it against the record and a mark can
    never be argued with.
    """

    start: int
    end: int
    #: The record's own name for the column, raw tag where it has one.
    attribute: str
    #: What the record types there. Kept whole, because a rejection is only readable with it.
    value: str


@dataclass(frozen=True)
class VariantSpan:
    """Where the near-identical records this document stands for DISAGREE — a half-open range.

    A **different kind of mark from** :class:`DeclaredSpan`, and the difference is load-bearing enough
    that they may never share a type. A declared span says *this quote is a column, refuse it*
    (:func:`~seqforge.harvest.verify._typed_column_at`). A variant span says *this quote speaks only
    for the member we sent, so do not fan it*. Conflate them and a fanned claim would be silently
    rejected, or — far worse — a quote lying inside a column would be fanned into 1439 records that
    never carried it. One field on the document per question asked of a span.

    ``value`` is the **exemplar's own** token text at this position, because a mark is only auditable
    beside the bytes it marks; ``n_values`` is how many distinct strings the group's members carry
    here, which is what makes "this is an index, not a fact" visible at all (it is undecidable from
    one record — :func:`~seqforge.harvest.plan.plan_extraction`'s collapse is what computes it).
    """

    start: int
    end: int
    #: The exemplar's token text in ``[start, end)``. Whole tokens, never a character slice.
    value: str
    #: Distinct strings the group's members carry at this position. ``>= 2`` by construction.
    n_values: int


@dataclass(frozen=True)
class NormalizedDoc:
    """One document reduced to the canonical span space, with both identities recorded.

    ``doc_sha256`` identifies the SOURCE bytes (stable document identity, what an Assertion cites);
    ``normalized_sha256`` identifies the span space itself, so a normalization drift is detectable.

    ``role`` is what the document IS to us, and it is **not** a property of the bytes: the same PDF is
    a reference when you cite it and an instruction when you write it for us. So it is set by the CLI
    from the flag the document arrived under, and it deliberately does NOT enter ``doc_sha256`` —
    otherwise one file would have two identities and its cached normalization would fork.

    ``scope``/``subject`` are what the document is ABOUT, and they carry the same disclaimer for the
    same reason: code sets them because code chose which record to render, and a document's contents
    never get a vote. ``subject`` is the record's accession at a record scope, and ``None`` for a
    document about the whole dataset. Together they are what lets a claim name a sample while
    ``AssertionDraft`` stays four fields wide.
    """

    doc_sha256: str
    normalized_sha256: str
    text: str
    source_basename: str
    role: DocRole = "reference"
    scope: DocScope = "dataset"
    subject: str | None = None
    normalizer_version: str = NORMALIZER_VERSION
    n_chars: int = 0
    #: Per-page ranges into ``text`` — non-empty only for a PDF, so a span can be tagged with its page.
    pages: tuple[PageSpan, ...] = ()
    #: Ranges of ``text`` that are this record repeating one of its own typed columns
    #: (:func:`declared_spans`). Empty for every document a human handed us, which have no columns.
    declared: tuple[DeclaredSpan, ...] = ()
    #: Ranges of ``text`` where the near-identical records this document stands for DISAGREE
    #: (:func:`variant_spans`). Empty unless the planner collapsed a group onto this document, and
    #: its complement — everything else in ``text`` — is byte-identical in every one of them, which
    #: is what lets a claim quoting only that complement be fanned to all of them.
    variants: tuple[VariantSpan, ...] = ()
    #: Which extractor produced ``text``: a :data:`PdfBackend` for a PDF, ``"text"`` otherwise.
    extractor: str = "text"


def normalize_text(raw: str) -> str:
    """Reduce prose to the canonical form quotes are matched against.

    Order matters: scrub characters that are not text at all first (a lone surrogate would crash the
    content-hash a few lines downstream), then kill soft hyphens and rejoin hyphen-broken words
    *before* collapsing newlines, or the line break that proves a word was split is already gone.
    """
    text = clean_invalid_unicode(raw)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    for src, dst in _LIGATURES.items():
        text = text.replace(src, dst)
    for src, dst in _PUNCT.items():
        text = text.replace(src, dst)
    text = unicodedata.normalize("NFKC", text)
    text = _WRAP_HYPHEN.sub("", text)  # "chemi-\nstry" -> "chemistry"
    text = _KEPT_HYPHEN.sub("-", text)  # "3-\nprime"    -> "3-prime" (hyphen is meaningful)
    text = _PARA_BREAK.sub(_PARA_TOKEN, text)  # protect real paragraph boundaries
    text = _LINE_BREAK.sub(" ", text)  # unwrap mid-sentence line breaks
    text = text.replace(_PARA_TOKEN, "\n\n")
    text = _WS_RUN.sub(" ", text)
    return "\n".join(line.strip() for line in text.split("\n")).strip()


#: Rows rendered per sheet before we stop and say so. A real sample table is tens to low hundreds of
#: rows; anything past this is a data dump (raw counts, a barcode list) beside the metadata, and this
#: text becomes an LLM prompt. The cap is PER SHEET, not per workbook, on purpose: a global budget
#: would let a giant junk sheet 0 starve the sample sheet 3 of room, which is exactly the dirty case.
_MAX_ROWS_PER_SHEET = 500


def _clean_cell(value: object) -> str:
    """One cell -> one contiguous token run.

    A cell is dirty in ways a row layout cannot survive: it may be ``None`` (blank, or a merged cell's
    hidden child), or it may hold its own newlines (``"treated with\\nDMSO"``). Collapsing every
    internal whitespace run to a single space keeps one cell to one token run, so an embedded newline
    can neither shatter the ` | ` columns nor fake a paragraph break in the canonical text.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _read_xlsx(path: Path) -> str:
    """Render EVERY sheet of a workbook to text — because the useful sheet is rarely the first one,
    and the workbook is rarely clean.

    A supplementary `.xlsx` is the common shape of a paper's sample metadata, and it is *plural* and
    *dirty*: the sheet that names the samples ("Sample metadata", "Experimental design") is usually not
    sheet 0, and it sits beside legends, notes, and dumped data. So the one thing this must not do is
    read a single sheet — which is exactly what `pandas.read_excel` does by default (`sheet_name=0`),
    and it would silently drop the sheet we came for. Every sheet is emitted, headed by its name so the
    model can see the book's shape.

    We deliberately do NOT try to find "the" metadata sheet or drop the irrelevant ones: that is
    interpretation, and a wrong guess loses data with no trace. Rendering all of it is cheap, the model
    reads the sheet that matters, and span verification means a dirty cell cannot reach the manifest
    without a quote that greps back to it. Code's only jobs here are mechanical — TRANSCRIBE faithfully
    (no header inference, no dtype coercion — `archive.py`'s discipline), keep each row its own
    paragraph so it survives `normalize_text` (a lone newline is a wrap artifact it flattens; a blank
    line is a boundary it keeps), and bound the size of a dumped sheet **visibly** (`_MAX_ROWS_PER_SHEET`).
    The rendering is deterministic — sheet order, then row/column order — because these bytes are what a
    quote is span-verified against, exactly as for `render_record`.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        blocks: list[str] = []
        for ws in wb.worksheets:
            rows = [f"Sheet: {ws.title}"]
            emitted = 0
            truncated = False
            for row in ws.iter_rows(values_only=True):
                cells = [_clean_cell(v) for v in row]
                while cells and not cells[-1]:
                    cells.pop()  # trailing empty cells carry nothing to quote
                if not any(cells):
                    continue  # a blank / spacer row is not a paragraph
                if emitted >= _MAX_ROWS_PER_SHEET:
                    truncated = True
                    break
                rows.append(" | ".join(cells))
                emitted += 1
            if truncated:
                # A marked truncation, never a silent one — visible to the model and to a human reading
                # the stored document, so "the table looked short" can never be mistaken for the data.
                rows.append(
                    f"[... more rows in sheet {ws.title!r} omitted after {_MAX_ROWS_PER_SHEET} ...]"
                )
            blocks.append("\n\n".join(rows))
    finally:
        wb.close()
    return "\n\n".join(blocks)


#: A single PDF page that extracts to more than this many characters is not a page of prose — it is a
#: bad read (a font/encoding failure dumping glyph tables, or an embedded data blob). Refuse it rather
#: than turn megabytes of noise into an LLM prompt: the PDF-text analogue of the FASTQ read budget,
#: where size is the bound and wall-clock never is.
_MAX_PAGE_CHARS = 1_280_000

#: Below this length a document is too short to judge by character entropy without false positives — a
#: terse instruction ("use GeneFull") is legitimately low-variety. Above it, text that is almost all
#: one repeated glyph is a garbled/scanned read with nothing to quote.
_MIN_CHARS_FOR_ENTROPY = 400
_MIN_ENTROPY = 2.5

#: Table rows rendered per page before we stop and say so — the same discipline as
#: ``_MAX_ROWS_PER_SHEET`` for a workbook: a table becomes an LLM prompt, and a run of hundreds of
#: rows is a data dump beside the metadata, not the metadata.
_MAX_TABLE_ROWS_PER_PAGE = 200


class UnreadableDocument(RuntimeError):
    """A source file produced no usable text — a scanned/encrypted PDF, or a garbled read.

    Raised by :func:`normalize_document` so the failure surfaces at the document boundary with a clear
    cause, instead of downstream when a truthful-looking quote fails to grep back and the blame lands
    on the hallucination tripwire. Fail loud and early: the extractor found nothing to quote, and no
    amount of LLM effort recovers a citation from bytes that are not there. It subclasses
    ``RuntimeError`` so the CLI's existing document-read handler already turns it into a nonzero exit.
    """


def _char_entropy(text: str) -> float:
    """Shannon entropy (bits) over the character distribution, spaces removed. Low ⇒ repetitive."""
    stripped = text.replace(" ", "")
    if not stripped:
        return 0.0
    n = len(stripped)
    return -sum((c / n) * math.log2(c / n) for c in Counter(stripped).values())


def _assess_readable(text: str, *, source: str) -> None:
    """Refuse a document with no quotable text. Empty always; near-zero entropy only past a length
    floor, so a short legitimate instruction is never mistaken for garble."""
    if not text.strip():
        raise UnreadableDocument(
            f"{source}: no text could be extracted (a scanned or image-only PDF?). Provide a "
            f"text-based PDF, or hand the prose as .txt/.md."
        )
    if len(text) >= _MIN_CHARS_FOR_ENTROPY and _char_entropy(text) < _MIN_ENTROPY:
        raise UnreadableDocument(
            f"{source}: extracted text is repetitive noise (entropy below {_MIN_ENTROPY}) — the PDF "
            f"likely failed to decode. Provide a text-based PDF, or hand the prose as .txt/.md."
        )


def _pypdf_pages(path: Path) -> list[str]:
    """Per-page text via pypdf (BSD). Content-stream order — no geometric reordering; see ``pymupdf``."""
    from pypdf import PdfReader

    return [page.extract_text() or "" for page in PdfReader(str(path)).pages]


def _pymupdf_pages(path: Path) -> list[str]:
    """Per-page text via PyMuPDF, in the PDF's own content-stream order (``sort=False``).

    PyMuPDF earns its place as the default for **robustness**, not clever ordering: on the worm corpus
    it read every manuscript, including two whose font tables make pypdf raise ``PdfReadError``. We
    deliberately do NOT pass ``sort=True``: geometric top-to-bottom sorting interleaves a two-column
    page line by line ("identified 118 Neurons share many common functions ... distinct classes among
    the 302 neurons"), because left- and right-column lines share a y and sort together. Content order,
    which a competent typesetter emits one full column at a time, reads cleanly — confirmed by an
    end-to-end harvest eval over the ten C. elegans papers.

    AGPL-3.0, and a declared dependency, so the import does not fail on a supported install. The
    ImportError branch is a clear refusal for a stripped-down environment, not a traceback.
    """
    try:
        import pymupdf
    except ImportError as exc:  # pragma: no cover - only on an install missing the dependency
        raise UnreadableDocument(
            "the 'pymupdf' PDF backend needs PyMuPDF (AGPL-3.0), a declared dependency that appears to "
            "be missing. Reinstall seqforge, or use `--pdf-backend pypdf` (BSD, no extra needed)."
        ) from exc
    with pymupdf.open(str(path)) as doc:
        return [page.get_text("text") for page in doc]


def _pdf_tables(path: Path) -> dict[int, list[str]]:
    """Per physical page (1-indexed), each detected table rendered as ` | `-joined rows.

    pdfplumber (MIT) detects tables and gives cells; we render them with the SAME discipline as a
    workbook sheet (:func:`_read_xlsx`): one cell to one contiguous token run (:func:`_clean_cell`),
    one row to one paragraph so ``normalize_text`` keeps rows apart, and a visible cap on a runaway
    table. The markdown is spliced INTO the page text by the caller — not held in a side channel — so a
    quoted cell greps back like any other span. A table's cells may also appear, scrambled, in the flat
    text; that duplication is faithful and harmless, and the rendered rows are the copy a quote matches.
    """
    import pdfplumber

    out: dict[int, list[str]] = {}
    with pdfplumber.open(str(path)) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            blocks: list[str] = []
            for table in page.find_tables():
                rows: list[str] = []
                truncated = False
                for row in table.extract():
                    if len(rows) >= _MAX_TABLE_ROWS_PER_PAGE:
                        truncated = True
                        break
                    cells = [_clean_cell(c) for c in row]
                    while cells and not cells[-1]:
                        cells.pop()  # trailing empty cells carry nothing to quote
                    if not any(cells):
                        continue  # a blank / spacer row is not a paragraph
                    rows.append(" | ".join(cells))
                if truncated:
                    rows.append(
                        f"[... more rows in this table omitted after {_MAX_TABLE_ROWS_PER_PAGE} ...]"
                    )
                if rows:
                    blocks.append("\n\n".join(rows))
            if blocks:
                out[i] = blocks
    return out


def _pdf_raw_pages(path: Path, *, backend: PdfBackend) -> list[str]:
    """Per-page raw text for a PDF: reading-order text from ``backend``, with that page's tables
    rendered to markdown and appended. Page count/order are the text backend's; tables are always
    pdfplumber's (core), so the backend choice only decides reading order, never whether tables appear.
    """
    text_pages = _pypdf_pages(path) if backend == "pypdf" else _pymupdf_pages(path)
    tables = _pdf_tables(path)
    pages: list[str] = []
    for i, page_text in enumerate(text_pages, start=1):
        if len(page_text) > _MAX_PAGE_CHARS:
            raise UnreadableDocument(
                f"{path.name}: page {i} extracted to {len(page_text)} characters, over the "
                f"{_MAX_PAGE_CHARS} limit — a decode failure, not a page of prose."
            )
        block = page_text
        if i in tables:
            rendered = "\n\n".join(tables[i])
            table_block = f"Table (page {i}):\n\n{rendered}"
            block = f"{block}\n\n{table_block}" if block.strip() else table_block
        pages.append(block)
    return pages


def _normalize_pages(raw_pages: list[str]) -> tuple[str, tuple[PageSpan, ...]]:
    """Normalize each page independently, join with a paragraph break, and record where each landed.

    Per-page (not whole-document) normalization is what makes the offset→page map exact, and it is
    behaviour-preserving: pages were already joined by a blank line, which ``normalize_text`` treats as
    a paragraph boundary and never rejoins across — so nothing that used to be one word is split. A
    page that normalizes to nothing contributes no span and no text.
    """
    text = ""
    spans: list[PageSpan] = []
    for i, raw in enumerate(raw_pages, start=1):
        norm = normalize_text(raw)
        if not norm:
            continue
        if text:
            text += "\n\n"
        start = len(text)
        text += norm
        spans.append(PageSpan(number=i, start=start, end=len(text)))
    return text, tuple(spans)


def page_for_offset(pages: tuple[PageSpan, ...], offset: int) -> int | None:
    """The 1-indexed page whose ``[start, end)`` contains ``offset`` — ``None`` if unpaged or between
    pages. Code-owned, mirroring how ``verify`` computes ``char_start``/``char_end`` from the quote."""
    if not pages:
        return None
    idx = bisect.bisect_right([pg.start for pg in pages], offset) - 1
    if 0 <= idx < len(pages) and pages[idx].start <= offset < pages[idx].end:
        return pages[idx].number
    return None


def read_document(path: Path) -> str:
    """Read a source document to raw text.

    PDF and XLSX are *extractors* behind the canonical-text contract, not special kinds of input:
    anything else falls through to plain text, so a hand-written `.md` or a `.txt` works with no extra
    code. The contract is the load-bearing part — `normalize_text` produces the one canonical string
    that span verification greps against, whatever the source format was. An `.xlsx` is a zip of XML,
    so it MUST take the extractor branch: read as plain text it would be replacement-char garbage, and
    a truthful quote could never grep back.

    `pypdf`/`openpyxl`/`pdfplumber` are imported lazily because these are the uncommon cases, but all
    are **declared dependencies** now, so the import does not fail. pypdf was once undeclared, with a
    remedy telling the user to install it by hand — which meant no supported install of seqforge could
    read a paper, the one document type the pilot dataset actually ships; openpyxl and pdfplumber carry
    the same weight for the supplementary tables papers ship their sample metadata in.

    This returns flat text via the default backend (:data:`DEFAULT_PDF_BACKEND`), with tables spliced
    in. :func:`normalize_document` owns the richer, page-aware PDF path — it is what tracks per-page
    offsets and honours ``--pdf-backend`` — so this stays the simple raw reader the non-PDF formats
    also use.
    """
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return "\n\n".join(_pdf_raw_pages(path, backend=DEFAULT_PDF_BACKEND))
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    return path.read_text(encoding="utf-8", errors="replace")


def normalize_document(
    path: str | Path, *, role: DocRole = "reference", pdf_backend: PdfBackend = DEFAULT_PDF_BACKEND
) -> NormalizedDoc:
    """Turn one source document into its :class:`NormalizedDoc` (the canonical span space).

    ``role`` comes from the caller — i.e. from which CLI flag the document arrived under — because it
    is a fact about how the document was OFFERED, not about its contents. Never infer it from the
    filename: that would be spoofable by renaming a downloaded PDF.

    ``pdf_backend`` selects the PDF reading-order extractor (``"pypdf"`` default, ``"pymupdf"`` for the
    two-column fix); it is ignored for every other format. A PDF is normalized **per page**, so the
    result also carries a page index (:attr:`NormalizedDoc.pages`) that lets a span be tagged with its
    page — the canonical text is still the ordered page concatenation, so a quote greps back unchanged.
    Refuses a document that yields no quotable text (:class:`UnreadableDocument`), so a scanned or
    garbled PDF fails at the boundary rather than as a mysterious span-verification miss downstream.
    """
    p = Path(path)
    source_bytes = p.read_bytes()
    if p.suffix.lower() == ".pdf":
        try:
            text, pages = _normalize_pages(_pdf_raw_pages(p, backend=pdf_backend))
        except UnreadableDocument:
            raise  # an empty/missing-backend refusal is already the right shape
        except Exception as exc:
            # A parser choking on a malformed PDF (pypdf's PdfReadError on a bad font table, a mupdf
            # decode error) is a read failure, not a crash to leak as a traceback — real manuscripts
            # trip this. Refuse cleanly, and point at the other backend, which often survives what one
            # engine cannot parse.
            other = "pymupdf" if pdf_backend == "pypdf" else "pypdf"
            raise UnreadableDocument(
                f"{p.name}: the {pdf_backend} backend could not parse this PDF "
                f"({type(exc).__name__}: {exc}). Try `--pdf-backend {other}`."
            ) from exc
        extractor: str = pdf_backend
    else:
        text = normalize_text(read_document(p))
        pages = ()
        extractor = "text"
    _assess_readable(text, source=p.name)
    return NormalizedDoc(
        doc_sha256=hashlib.sha256(source_bytes).hexdigest(),
        normalized_sha256=hashlib.sha256(text.encode()).hexdigest(),
        text=text,
        source_basename=p.name,
        role=role,
        # A document a human handed us is about the whole pile of files. There is no other honest
        # reading of "here is the paper" — see `resolve/records.py` for what that costs a sample claim.
        scope="dataset",
        pages=pages,
        extractor=extractor,
        n_chars=len(text),
    )


def render_record(record: ArchiveRecord) -> str:
    """One archive record -> the text a model reads. Deterministic, and the ONLY renderer.

    Deterministic matters more than it looks. This text *is* the document: its sha256 is the identity
    an assertion cites, and the span check greps this exact string. So the rendering must be
    reproducible from the record forever — a human handed the record and this function must be able to
    regenerate the bytes a quote was checked against, or the citation is unfalsifiable.

    Only free text is rendered. The structured half (``strain = CQ758``) is code's to read and is
    already a key and a value, so putting it in front of a model would be asking it to transcribe
    something we can copy — a chance to be wrong and no chance to be useful.
    """
    lines = [f"{record.level} {record.accession}"]
    for ft in record.free_text:
        lines.append(f"{ft.label}: {ft.text}")
    return "\n\n".join(lines)


def _is_token_char(ch: str) -> bool:
    return ch.isalnum() or ch == "_"


def _on_token_boundary(text: str, start: int, end: int) -> bool:
    """Does ``text[start:end]`` stand as its own run of tokens rather than sit inside a longer one?

    The ``CD4``-inside-``CD45`` caution, one layer earlier than the resolver's: a typed value that
    happens to be spelled inside a longer word is a coincidence of characters, and marking it would
    claim the record re-rendered a column where it did nothing of the kind. Only the two edges are
    checked, and only where both sides of an edge are token characters — a value that begins or ends
    in punctuation (``3'``) needs no boundary there and must not be denied one.
    """
    if start > 0 and _is_token_char(text[start - 1]) and _is_token_char(text[start]):
        return False
    return not (end < len(text) and _is_token_char(text[end - 1]) and _is_token_char(text[end]))


def token_spans(text: str) -> tuple[tuple[int, int], ...]:
    """Every maximal run of token characters in ``text``, as half-open ``[start, end)`` ranges.

    The same :func:`_is_token_char` the boundary check uses, so ``_`` counts and ``nasal_prox1_270``
    is ONE token. That is not a detail to tune: see :func:`varying_token_indices` for what a
    finer-grained answer costs.
    """
    spans: list[tuple[int, int]] = []
    i, n = 0, len(text)
    while i < n:
        if not _is_token_char(text[i]):
            i += 1
            continue
        j = i
        while j < n and _is_token_char(text[j]):
            j += 1
        spans.append((i, j))
        i = j
    return tuple(spans)


def token_values(text: str) -> tuple[str, ...]:
    """``text``'s tokens, in order — what :func:`varying_token_indices` compares."""
    return tuple(text[s:e] for s, e in token_spans(text))


def token_skeleton(text: str) -> tuple[str, ...]:
    """Everything in ``text`` that is NOT a token: the separators, in order, leading and trailing ones
    included. Two texts with the same skeleton have the same token count and the same punctuation, so
    they differ only in what their tokens SAY — which is the only shape a group can be aligned in
    without parsing the prose.

    It is a whole-string equality key, deliberately. An edit-distance alignment would let two records
    with different fields be "near-identical" at some threshold, and a threshold is a knob nobody can
    tune from bytes; a record that structurally differs simply forms its own group and is asked on its
    own, which is the failure mode we want (:mod:`seqforge.harvest.plan`).
    """
    spans = token_spans(text)
    seps: list[str] = []
    prev = 0
    for start, end in spans:
        seps.append(text[prev:start])
        prev = end
    seps.append(text[prev:])
    return tuple(seps)


def varying_token_indices(texts: Sequence[str]) -> tuple[int, ...]:
    """Token positions where at least two of ``texts`` disagree. THE definition of "varies".

    **Token boundaries, never character spans, and this costs the last order of magnitude.** At
    character granularity ``age: 3`` in 1439 records and ``age: 30`` in one makes ``3`` a shared span:
    ``age = 3`` would fan into the record that says 30, and the leftover variant would be the single
    character ``0``, which no askable-content test would ever send — a wrong claim, silently, on the
    one record that differed. It is the ``CD4``-inside-``CD45`` caution (:func:`_on_token_boundary`)
    one layer up, and it is why ``3`` and ``30`` are two tokens here and never a shared prefix.

    Two rescues for the granularity this costs were considered and both rejected. Masking a digit run
    whose values across the group are exactly an enumeration ``1..N`` works on a plate and dies on
    ``sample_1h ... sample_24h``, where the digit IS the data and no code-side test separates the two.
    And ``_is_token_char`` counting ``_`` is not an oversight to fix: it makes ``nasal_prox1_270`` one
    token, so a serially-named deposit's sample documents do not collapse and every one of them is
    asked. That is the design pricing itself, not a defect — cost tracks information content.

    Raises ``ValueError`` unless every text shares one skeleton; a caller that has not grouped by
    :func:`token_skeleton` first is asking to align two different documents.
    """
    if not texts:
        return ()
    skeleton = token_skeleton(texts[0])
    values = [token_values(t) for t in texts]
    for text in texts[1:]:
        if token_skeleton(text) != skeleton:
            raise ValueError("varying_token_indices needs one skeleton; group before you align")
    return tuple(k for k in range(len(values[0])) if len({v[k] for v in values}) > 1)


def variant_spans(texts: Sequence[str]) -> tuple[VariantSpan, ...]:
    """``texts[0]``'s ranges at every position :func:`varying_token_indices` names.

    Marks, never splices. The bytes sent stay one member's own rendering, so ``doc_sha256`` is still
    the sha of a string :func:`render_record` regenerates from one record, and the model still reads
    coherent prose rather than a concatenation of fragments — a fragment document would be a string no
    record ever produced, and a quote into it could be checked against nothing.
    """
    if len(texts) < 2:
        return ()
    varying = varying_token_indices(texts)
    spans = token_spans(texts[0])
    values = [token_values(t) for t in texts]
    return tuple(
        VariantSpan(
            start=spans[k][0],
            end=spans[k][1],
            value=texts[0][spans[k][0] : spans[k][1]],
            n_values=len({v[k] for v in values}),
        )
        for k in varying
    )


def variant_text(text: str, indices: Sequence[int]) -> str:
    """``text`` reduced to the tokens at ``indices`` — its labels and its shared prose dropped.

    What a **non-exemplar** member of a collapsed group is sent as. The exemplar carries the group's
    prose in full, marked (:func:`variant_spans`); sending 1439 more copies of the identical paragraph
    to read five serial names is the bill this whole mechanism exists to stop. Measured on the
    1440-record GSE207085 dump: an experiment record renders to 429 characters of which 67 vary, and
    the shared 362 are byte-identical in all 1440.

    **The record's own accession leads, because it varies and everything that varies is kept.** No
    special case, and it is what makes the reduction safe: a document's identity is its bytes, so two
    members that happen to share a serial name would otherwise reduce to one ``doc_sha256`` — and
    ``resolve`` keys a document's subject by that sha, so one of the two samples would silently
    inherit the other's claims. It also keeps ``_nothing_to_ask`` a single sentence: a member is
    withheld exactly when this text would say nothing but its own name.

    **Adjacent varying tokens keep the separator the record wrote between them; anything else gets a
    paragraph break.** Adjacency is the only syntax the group did NOT already read in the exemplar, so
    it is the only syntax worth carrying; and a paragraph break between tokens that were apart stops
    the join from inventing a phrase no record wrote (``genotype: WT`` and ``age: 72`` must not
    reduce to ``WT 72``). What is genuinely lost is the syntax BETWEEN two variants — ``5 mM DMSO``
    with a shared ``mM`` reduces to ``5`` and ``DMSO`` — and that loss is bounded by the two checks
    that do not move: a quote must still grep back into this text, and it must still entail its value.
    """
    spans = token_spans(text)
    out: list[str] = []
    previous: int | None = None
    for k in indices:
        if previous is not None:
            joiner = text[spans[previous][1] : spans[k][0]] if k == previous + 1 else "\n\n"
            out.append(joiner)
        out.append(text[spans[k][0] : spans[k][1]])
        previous = k
    return normalize_text("".join(out))


def is_invariant_span(doc: NormalizedDoc, start: int, end: int) -> bool:
    """Does ``doc.text[start:end]`` lie wholly inside text byte-identical across ``doc``'s group?

    The fan-out predicate, and it is stated as *touches no variant* rather than *is inside a shared
    range* because those are the same set and only one of them can be got wrong: a claim is fanned
    exactly when nothing it quotes is a place the members disagree. A document with no marks has no
    group, so this is vacuously true and the caller — never this function — is what knows whether
    there is anything to fan to.
    """
    return not any(v.start < end and start < v.end for v in doc.variants)


def declared_spans(text: str, records: Sequence[ArchiveRecord]) -> tuple[DeclaredSpan, ...]:
    """Every span of ``text`` byte-equal to a value one of ``records`` typed into a column of its own.

    A record carries a fact twice: once as a structured column code already reads, and once inside
    the prose it renders for a human. SRA's experiment title is the shape that measured this — an
    experiment types ``library_strategy = "RNA-Seq"`` and then ends its title with the same word, so a
    model asked for the chemistry answers with it, quoting verbatim. The quote greps back, entailment
    is vacuous because the value sits inside its own quote, and a transcription of a column arrives
    looking exactly like a reading of prose (#184).

    Marking it is the only way to tell those apart, and the mark has to be **provable**: a span is
    marked when it is byte-equal to a value on the record that produced this text, whole tokens only.
    Nothing here parses the prose. There is no delimiter grammar, no accession format, no notion of
    which archive rendered the string — #188's "split the title on ``;``" is exactly what this must
    not be, because seqforge compiles in-house datasets that never had an accession and other
    archives lay their records out differently. Where there are no records there are no columns, and
    this is a no-op that returns ``()``.

    ``records`` is plural because a document may be the concatenation of several records' renderings
    (:func:`~seqforge.harvest.plan._collapsed_run_document` folds one sample's runs into one). They
    are levels of one deposit, so a span byte-equal to any member's column is still that submitter
    repeating themselves; offsets are computed against the joined text, which is the only string a
    quote is ever checked against.
    """
    found: list[DeclaredSpan] = []
    for record in records:
        for attr in record.attributes:
            # Fold the needle the same way the haystack was folded, or a value the archive stored
            # with a non-breaking space could never match its own rendering.
            needle = normalize_text(attr.value)
            if not needle:
                continue
            start = text.find(needle)
            while start >= 0:
                end = start + len(needle)
                if _on_token_boundary(text, start, end):
                    found.append(
                        DeclaredSpan(
                            start=start,
                            end=end,
                            attribute=attr.raw_name or attr.name,
                            value=attr.value,
                        )
                    )
                start = text.find(needle, start + 1)
    # One span, one mark. Two runs of a sample type the same `library_strategy`, and each of them
    # finds it in both of their renderings, so the same range would otherwise be marked once per
    # (record x occurrence) — four copies of two spans, and any count over them a fiction. Which
    # column proves a span is a detail; that it is proven at all is the fact.
    unique = {
        (d.start, d.end): d for d in sorted(found, key=lambda d: (d.start, d.end, d.attribute))
    }
    return tuple(unique.values())


def normalize_record(record: ArchiveRecord) -> NormalizedDoc:
    """An archive record -> its own document, scoped to itself.

    This is the whole mechanism behind "a claim can name a sample without being able to name a
    sample". The document holds one record's prose, so whatever the model finds in it is about that
    record, because that is the only thing in it. ``subject`` is set from the record we rendered —
    code knows it because code chose it, exactly as ``instruct.py`` decides document role.

    It is also where the record's own columns are marked in its own text (:func:`declared_spans`), so
    that a quote which merely repeats one can be told from a reading of the prose around it.
    """
    text = normalize_text(render_record(record))
    digest = hashlib.sha256(text.encode()).hexdigest()
    return NormalizedDoc(
        # The rendering IS the source: there are no other bytes to identify. Both hashes are the same
        # string's, and saying so is more honest than inventing a second identity for one document.
        doc_sha256=digest,
        normalized_sha256=digest,
        text=text,
        source_basename=f"{record.level}-{record.accession}.txt",
        role="reference",  # an archive record is a database field, never an instruction to us
        scope=record.level,
        subject=record.accession,
        n_chars=len(text),
        declared=declared_spans(text, [record]),
    )


def has_prose(record: ArchiveRecord) -> bool:
    """Is there anything here for a model to read? A record with no free text is not worth a call."""
    return any(ft.text.strip() for ft in record.free_text)
