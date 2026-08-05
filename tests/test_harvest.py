"""Tests for ``harvest``: the canonical span space and the hallucination tripwire.

The adversarial cases ARE the feature. A tripwire that only passes honest input proves nothing — so
these assert that fabricated provenance and a real-quote-wrong-value both get rejected, and equally
that a truthful quote mangled by PDF wrapping is NOT rejected (a tripwire with false positives is one
we would soon learn to ignore).
"""

from __future__ import annotations

from pathlib import Path

import pytest

from seqforge.harvest import (
    DEFAULT_PDF_BACKEND,
    NormalizedDoc,
    PdfBackend,
    UnreadableDocument,
    clean_invalid_unicode,
    entails,
    find_span,
    normalize_document,
    normalize_text,
    verify_drafts,
)
from seqforge.models.assertion import AssertionDraft, ExtractorProvenance, SourceSpan
from seqforge.models.records import ArchiveRecord, FreeText, RecordAttribute

EXTRACTOR = ExtractorProvenance(model_id="test-model", prompt_version="v1")


def _esc(s: str) -> str:
    return s.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")


def _make_pdf(pages: list[str]) -> bytes:
    """A minimal, valid multi-page PDF built from raw operators — no dependency, no binary fixture.

    Each entry in ``pages`` is one page's text (newline-separated lines). Enough to exercise the real
    pypdf/pymupdf extractors, page tagging, and the empty-document tripwire without shipping a blob.
    """
    objs: dict[int, bytes] = {3: b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>"}
    kids: list[int] = []
    num = 4
    for page in pages:
        lines = ["BT", "/F1 11 Tf"]
        y = 720
        for ln in page.split("\n"):
            lines.append(f"1 0 0 1 72 {y} Tm ({_esc(ln)}) Tj")
            y -= 16
        lines.append("ET")
        content = "\n".join(lines).encode()
        cnum, pnum = num, num + 1
        num += 2
        objs[cnum] = b"<</Length %d>>\nstream\n%s\nendstream" % (len(content), content)
        objs[pnum] = (
            b"<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
            b"/Resources<</Font<</F1 3 0 R>>>>/Contents %d 0 R>>" % cnum
        )
        kids.append(pnum)
    objs[1] = b"<</Type/Catalog/Pages 2 0 R>>"
    objs[2] = b"<</Type/Pages/Kids[%s]/Count %d>>" % (
        b" ".join(b"%d 0 R" % k for k in kids),
        len(kids),
    )
    out = bytearray(b"%PDF-1.4\n")
    offsets: dict[int, int] = {}
    for i in range(1, num):
        offsets[i] = len(out)
        out += b"%d 0 obj\n%s\nendobj\n" % (i, objs[i])
    xref = len(out)
    out += b"xref\n0 %d\n0000000000 65535 f \n" % num
    for i in range(1, num):
        out += b"%010d 00000 n \n" % offsets[i]
    out += b"trailer\n<</Size %d/Root 1 0 R>>\nstartxref\n%d\n%%%%EOF\n" % (num, xref)
    return bytes(out)


def _pdf(tmp_path: Path, pages: list[str], name: str = "paper.pdf") -> Path:
    p = tmp_path / name
    p.write_bytes(_make_pdf(pages))
    return p


# ---------- normalize: the span space ----------
def test_normalize_dehyphenates_across_a_line_wrap() -> None:
    # a PDF splits "chemistry" across lines; a truthful quote must still be findable
    assert "chemistry" in normalize_text("we used the chemi-\nstry described")


def test_normalize_unwraps_lines_but_keeps_paragraphs() -> None:
    out = normalize_text("first line\nsecond line\n\nnew paragraph")
    assert "first line second line" in out  # a lone newline was a wrap artifact
    assert "\n\n" in out  # a blank line was a real boundary


def test_normalize_expands_ligatures_and_smart_punctuation() -> None:
    out = normalize_text("the ﬁrst 3′ kit — “quoted” ‘v3’")
    assert "first" in out  # ﬁ ligature
    assert "3'" in out  # prime -> apostrophe
    assert '"quoted"' in out and "'v3'" in out  # smart quotes flattened
    assert "-" in out  # em dash -> hyphen


def test_normalize_strips_soft_hyphen_and_nbsp() -> None:
    out = normalize_text("single­cell RNA-seq")
    assert "singlecell RNA-seq" in out


def test_normalize_document_records_both_identities(tmp_path: Path) -> None:
    doc = tmp_path / "methods.txt"
    # deliberately un-canonical (ligature + wrap), so normalization actually does something
    doc.write_text("Libraries used the ﬁrst Chromium Single\nCell 3' v3 kit.")
    nd = normalize_document(doc)
    assert nd.source_basename == "methods.txt"
    assert len(nd.doc_sha256) == 64 and len(nd.normalized_sha256) == 64
    # source identity != span-space identity once normalization has changed anything
    assert nd.doc_sha256 != nd.normalized_sha256
    assert nd.n_chars == len(nd.text)
    assert "first Chromium Single Cell 3' v3" in nd.text


def test_a_multi_sheet_xlsx_renders_every_sheet_not_just_the_first(tmp_path: Path) -> None:
    """A supplementary `.xlsx` is plural, and the sample sheet is rarely sheet 0.

    The failure this guards against is silent: `pandas.read_excel` defaults to one sheet, so a naive
    reader would drop "Sample metadata" and harvest would find nothing to quote — no error, just an
    empty extraction. So the load-bearing assertion is that a value living ONLY on the second sheet
    survives into the canonical text, and that a quote of it verifies end to end.
    """
    from openpyxl import Workbook

    wb = Workbook()
    design = wb.active
    assert design is not None, "a fresh Workbook always has an active sheet"
    design.title = "Experimental design"
    design.append(["note", "single-cell RNA-seq, two conditions"])
    samples = wb.create_sheet("Sample metadata")
    samples.append(["sample_id", "strain", "genotype"])
    samples.append(["GSM001", "N2", "wild-type"])
    samples.append(["GSM002", "CB1370", "daf-2 mutant"])
    xlsx = tmp_path / "supplementary_tables.xlsx"
    wb.save(xlsx)

    nd = normalize_document(xlsx)
    # both sheets are present and labelled, so the model can see the book's shape...
    assert "Sheet: Experimental design" in nd.text
    assert "Sheet: Sample metadata" in nd.text
    # ...and a value that exists ONLY on the second sheet made it through (the whole point).
    assert "CB1370" in nd.text and "daf-2 mutant" in nd.text

    # a quote of a second-sheet cell greps back and entails its value — the round-trip works.
    draft = AssertionDraft(
        field="experiment.samples.genotype",
        value="wild-type",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="N2 | wild-type"),
        llm_confidence=0.9,
    )
    assert verify_drafts([draft], [nd], extractor=EXTRACTOR).n_accepted == 1


def test_a_dirty_xlsx_transcribes_faithfully_without_cleaning(tmp_path: Path) -> None:
    """A real supplementary workbook is messy: a title row, notes, blank spacer rows, a multi-line
    cell, and a sheet that has nothing to do with samples. We do NOT clean or drop any of it — we
    transcribe, and the model reads what it needs. The mechanical guarantees are that a value survives
    intact (an embedded newline does not shatter its row) and that an irrelevant sheet is harmless.
    """
    from openpyxl import Workbook

    wb = Workbook()
    legend = wb.active
    assert legend is not None, "a fresh Workbook always has an active sheet"
    legend.title = "Legend"  # an irrelevant sheet — rendered, not dropped, and it breaks nothing
    legend.append(["colour code", "red = failed QC"])
    samples = wb.create_sheet("Table S3")
    samples.append(
        ["Supplementary Table S3. Sample metadata"]
    )  # a title/preamble row, not a header
    samples.append([])  # a blank spacer row
    samples.append(["sample_id", "strain", "treatment"])
    samples.append(["GSM001", "N2", "heat shock\n34C for 2h"])  # a multi-line cell

    xlsx = tmp_path / "dirty.xlsx"
    wb.save(xlsx)
    nd = normalize_document(xlsx)

    assert "Sheet: Legend" in nd.text and "Sheet: Table S3" in nd.text
    # the multi-line cell collapsed to one contiguous run, so the row stayed one line and greps back
    assert "heat shock 34C for 2h" in nd.text
    draft = AssertionDraft(
        field="experiment.samples.treatment",
        value="heat shock",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="N2 | heat shock 34C for 2h"),
        llm_confidence=0.8,
    )
    assert verify_drafts([draft], [nd], extractor=EXTRACTOR).n_accepted == 1


def test_a_dumped_sheet_is_bounded_and_the_truncation_is_marked(tmp_path: Path) -> None:
    """A workbook may carry a raw-data sheet of thousands of rows beside the metadata. That text
    becomes an LLM prompt, so it is bounded per sheet — and, per "no silent caps", the cut is MARKED so
    a dropped tail can never be mistaken for the whole table.
    """
    from openpyxl import Workbook

    from seqforge.harvest.normalize import _MAX_ROWS_PER_SHEET

    wb = Workbook()
    dump = wb.active
    assert dump is not None, "a fresh Workbook always has an active sheet"
    dump.title = "counts"
    for i in range(_MAX_ROWS_PER_SHEET + 50):
        dump.append([f"gene_{i}", i])
    dump.append(["LAST_ROW_SENTINEL", -1])
    xlsx = tmp_path / "big.xlsx"
    wb.save(xlsx)
    nd = normalize_document(xlsx)

    assert "gene_0" in nd.text  # the head is kept
    assert "LAST_ROW_SENTINEL" not in nd.text  # the tail past the cap is dropped...
    assert "omitted after" in nd.text  # ...and the drop is announced, not silent


# ---------- find_span ----------
def test_find_span_tolerates_whitespace_differences() -> None:
    text = "Libraries were prepared with the Chromium Single Cell 3' v3 kit."
    assert find_span(text, "Chromium Single Cell 3' v3") is not None
    # the same quote as copied across a line wrap (extra/newline whitespace) still resolves
    span = find_span(text, "Chromium  Single\nCell 3' v3")
    assert span is not None and text[span[0] : span[1]].startswith("Chromium")
    assert find_span(text, "Chromium Single Cell 5' v2") is None


# ---------- entailment ----------
def test_entailment_accepts_a_kb_alias_carried_by_the_quote() -> None:
    # the paper never writes "10x-3p-gex-v3"; it writes the vendor's prose. KB aliases bridge that.
    assert entails("Chromium Single Cell 3' v3 kit", "library.chemistry", "10x-3p-gex-v3")


def test_entailment_rejects_a_real_quote_pinned_to_a_wrong_value() -> None:
    """The design's named failure: a verbatim span attached to a conclusion it does not support."""
    assert not entails("We performed single-cell RNA-seq.", "library.chemistry", "10x-3p-gex-v3")


def test_entailment_rejects_a_different_version() -> None:
    assert not entails("Chromium Single Cell 3' v2 kit", "library.chemistry", "10x-3p-gex-v3")


def test_entailment_accepts_a_form_that_only_describes_the_run() -> None:
    """A `descriptive_alias` is demoted in RANKING, and must stay whole here (#266).

    The two questions share `curated_forms` deliberately, and they are not the same question. Ranking
    asks *which node does this value name*, where "paired-end RNA-seq" must lose to any chemistry the
    value also names. Entailment asks *does this quote support a value already established to be
    `bulk-rnaseq`* — and there it does, because the node is no longer in doubt. Dropping the
    descriptive forms from `curated_forms` to fix the first question would silently answer the second
    one differently, which is the drift the shared list exists to prevent.
    """
    assert entails(
        "Illumina PE RNA-seq libraries were prepared", "library.chemistry", "bulk-rnaseq"
    )
    assert entails("paired-end RNA-seq of whole embryos", "library.chemistry", "bulk-rnaseq")
    # ...and demotion is not deletion in the other direction either: a named form still entails
    assert entails("a bulk RNA-seq library", "library.chemistry", "bulk-rnaseq")


def test_entailment_plain_value_substring() -> None:
    assert entails(
        "The organism was Caenorhabditis elegans.", "experiment.organism", "Caenorhabditis elegans"
    )


def test_entailment_is_vacuous_when_the_value_is_copied_out_of_the_quote() -> None:
    """Pin the LIMIT of span verification, so nobody mistakes it for a guarantee it does not offer.

    The check's power comes from `value` being a controlled-vocabulary term whose surface forms must
    appear in the quote. For a free-text field the model copies the value out of the quote, so the
    substring test is trivially satisfied and this returns True for ANY field label.

    Hence: entailment catches fabricated and mis-attributed claims, never field-assignment errors. A
    real quote filed under the wrong field passes here by construction — `eval run --llm` caught
    exactly that (standard worm husbandry filed as an experimental `condition`). The defense is the
    prompt's field definition plus the evals corpus, not this function.
    """
    husbandry = "maintained on NGM plates seeded with E. coli OP50 at 20 C"
    quote = f"Caenorhabditis elegans were {husbandry}."
    # A true statement, correctly quoted, filed under a field it does not belong in — and entails()
    # cannot tell. It would say yes to any field name at all:
    assert entails(quote, "experiment.samples.condition", husbandry)
    assert entails(quote, "experiment.samples.tissue", husbandry)
    assert entails(quote, "totally.made.up.field", husbandry)
    assert not entails(
        "The organism was Homo sapiens.", "experiment.organism", "Caenorhabditis elegans"
    )


# ---------- verify: the tripwire end to end ----------
def _doc(tmp_path: Path, text: str) -> NormalizedDoc:
    p = tmp_path / "methods.txt"
    p.write_text(text)
    return normalize_document(p)


def test_verify_accepts_a_truthful_span_and_computes_offsets(tmp_path: Path) -> None:
    nd = _doc(tmp_path, "Libraries were prepared with the Chromium Single Cell 3' v3 kit.")
    draft = AssertionDraft(
        field="library.chemistry",
        value="10x-3p-gex-v3",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="Chromium Single Cell 3' v3"),
        llm_confidence=0.9,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 1 and not report.rejected
    a = report.assertions[0]
    assert a.span_verified and a.entailment_ok
    # offsets are code-computed, and they really point at the quote
    assert nd.text[a.span.char_start : a.span.char_end] == "Chromium Single Cell 3' v3"


def test_verify_rejects_fabricated_provenance(tmp_path: Path) -> None:
    nd = _doc(tmp_path, "Libraries were prepared with the Chromium Single Cell 3' v3 kit.")
    draft = AssertionDraft(
        field="library.chemistry",
        value="10x-3p-gex-v3",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="the 10x Chromium v3 protocol (Figure 2)"),
        llm_confidence=0.99,  # high confidence must not rescue an invented quote
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 0
    assert report.rejected[0]["reason"] == "span_not_found"


def test_verify_rejects_real_quote_wrong_value(tmp_path: Path) -> None:
    nd = _doc(tmp_path, "We performed single-cell RNA-seq on dissociated tissue.")
    draft = AssertionDraft(
        field="library.chemistry",
        value="10x-3p-gex-v3.1",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="single-cell RNA-seq"),
        llm_confidence=0.95,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 0
    assert report.rejected[0]["reason"] == "not_entailed"


def test_verify_rejects_a_quote_citing_an_unknown_document(tmp_path: Path) -> None:
    nd = _doc(tmp_path, "Chromium Single Cell 3' v3 kit.")
    draft = AssertionDraft(
        field="library.chemistry",
        value="10x-3p-gex-v3",
        span=SourceSpan(doc_sha256="0" * 64, quote="Chromium Single Cell 3' v3"),
        llm_confidence=0.9,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 0 and report.rejected[0]["reason"] == "unknown_doc"


def test_verify_accepts_a_quote_broken_by_pdf_wrapping(tmp_path: Path) -> None:
    """A truthful quote mangled by PDF wrapping must survive — a tripwire with false positives is one
    we would soon learn to ignore. Exercises all three hazards at once: a SEMANTIC hyphen at a wrap
    (3-prime), a wrap hyphen inside a word (chemi-stry), and a plain mid-sentence line break."""
    nd = _doc(
        tmp_path,
        "Libraries used the Chromium Single Cell 3-\nprime v3 chemi-\nstry\nthroughout the study.",
    )
    assert (
        "3-prime v3 chemistry throughout" in nd.text
    )  # meaningful hyphen kept, wrap hyphen closed
    draft = AssertionDraft(
        field="library.chemistry",
        value="10x-3p-gex-v3",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="Chromium Single Cell 3-prime v3"),
        llm_confidence=0.8,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 1, report.rejected


# ---------- the field allowlist: the check span verification cannot stand in for ----------
def test_verify_rejects_a_field_nobody_authorized(tmp_path: Path) -> None:
    """A REAL quote, entailing a REAL value, on a field that was never on offer.

    This is not a hypothetical. `AssertionDraft.field` is a plain `str` (it must be, to stay inside
    every provider's strict-schema subset), and `DEFAULT_FIELDS` was only ever interpolated into the
    prompt — `verify` never compared a returned draft against it. Both span-verification checks pass
    here on their own terms: the quote is verbatim and it genuinely supports "10". Span verification
    asks "is this claim in the document?"; the question this draft needs is "may you set this field at
    all?", and only an allowlist can answer it. Without it, prose becomes aligner argv, which is
    precisely what we forbid.
    """
    nd = _doc(tmp_path, "For this dataset, add --outFilterMismatchNmax 10 to the alignment.")
    draft = AssertionDraft(
        field="processing.params.outFilterMismatchNmax",
        value="10",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="add --outFilterMismatchNmax 10"),
        llm_confidence=0.95,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 0
    assert report.rejected[0]["reason"] == "field_not_permitted"

    # ...and prove the rejection is the ALLOWLIST talking, not a weak quote: both span-verification checks pass.
    from seqforge.harvest.verify import entails, find_span

    assert find_span(nd.text, draft.span.quote) is not None
    assert entails(draft.span.quote, draft.field, draft.value)


def test_verify_still_accepts_every_permitted_field(tmp_path: Path) -> None:
    """The allowlist must not be so tight it rejects the fields we actually ask for."""
    from seqforge.harvest.fields import DEFAULT_FIELDS, PERMITTED_FIELDS

    assert set(DEFAULT_FIELDS) <= PERMITTED_FIELDS
    nd = _doc(tmp_path, "We profiled Caenorhabditis elegans neurons, deposited as PRJNA1027859.")
    draft = AssertionDraft(
        field="experiment.organism",
        value="Caenorhabditis elegans",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="Caenorhabditis elegans"),
        llm_confidence=0.9,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 1, report.rejected


# The allowlist's own truth table -- `is_permitted` as an exact match, not a prefix rule -- lives in
# `tests/test_fields.py`, beside the vocabulary it reads. What stays here is `verify` consulting it.


# ---------- doc role: only a document you hand us may steer the pipeline ----------
def test_a_reference_doc_may_not_set_processing(tmp_path: Path) -> None:
    """A downloaded methods PDF must never reach the aligner.

    The quote is real and it entails GeneFull — both span-verification checks pass. What rejects it is the document's
    ROLE, which code owns because code chose the flag. This is a deliberate narrowing of "instructions
    may live among the unstructured metadata", and it costs nothing: with the all-five default a
    paper saying "we used GeneFull" describes a subset of what we already compute.
    """
    from seqforge.harvest import normalize_document

    doc = tmp_path / "paper.txt"
    doc.write_text("Reads were quantified in GeneFull mode against ce11.")
    nd = normalize_document(doc)  # default role: reference
    assert nd.role == "reference"
    draft = AssertionDraft(
        field="processing.quantification",
        value="GeneFull",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="quantified in GeneFull mode"),
        llm_confidence=0.9,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 0
    assert report.rejected[0]["reason"] == "field_not_permitted_for_doc"

    # ...and the SAME bytes, offered as an instruction, are accepted. Role is not a property of the
    # file — it is a property of how it was offered.
    nd_i = normalize_document(doc, role="instruction")
    assert nd_i.doc_sha256 == nd.doc_sha256, "role must not fork a document's identity"
    assert verify_drafts([draft], [nd_i], extractor=EXTRACTOR).n_accepted == 1


def test_entailment_is_non_vacuous_for_a_closed_vocabulary_field(tmp_path: Path) -> None:
    """The one field where entailment actually bites — and it must stay that way.

    `entails` is vacuous when value ⊆ quote, so it does real work ONLY for a controlled vocabulary.
    soloFeatures is closed (six STARsolo values), so a quote must NAME the feature. A quote describing
    the biology does not, and rejecting it is the RIGHT answer: inferring "nuclei -> GeneFull" is an
    inference code owns, not the model. If surface_forms ever learned that alias, span verification would be theatre.
    """
    from seqforge.harvest.verify import entails

    assert entails("should be aligned in GeneFull mode", "processing.quantification", "GeneFull")
    assert entails("pass --soloFeatures GeneFull", "processing.quantification", "GeneFull")
    assert entails(
        "quantify with GeneFull_Ex50pAS", "processing.quantification", "GeneFull_Ex50pAS"
    )
    # ...and the line that must NOT be crossed:
    assert not entails("we prepared single nuclei", "processing.quantification", "GeneFull")
    assert not entails("count introns too", "processing.quantification", "GeneFull")
    assert not entails("this is a pre-mRNA rich sample", "processing.quantification", "GeneFull")


def test_entailment_of_prep_type_is_by_normalized_prep_not_verbatim_tokens() -> None:
    """`library.prep_type` is biology the model reports, not a STARsolo flag: a terse quote must support
    a verbose value, and a cell quote must never support a nucleus value. GSE229022 lost GeneFull-primary
    because the verbose value "single nucleus RNA sequencing (snRNA-seq)" was not entailed by a short
    "snRNA-seq" span under the generic token-subset matcher; the normalized-prep check fixes that.
    """
    from seqforge.harvest.verify import entails

    # a terse quote entails a verbose nucleus value (the GSE229022 case)
    assert entails(
        "libraries were profiled by snRNA-seq",
        "library.prep_type",
        "single nucleus RNA sequencing (snRNA-seq)",
    )
    assert entails("nuclei were isolated from frozen tissue", "library.prep_type", "single-nucleus")
    # symmetric: a cell quote entails a cell value
    assert entails("a single-cell suspension was loaded", "library.prep_type", "single-cell")
    # the entailment tripwire still bites: a cell quote must NOT support a nucleus value
    assert not entails("a single-cell suspension was loaded", "library.prep_type", "single-nucleus")
    # a quote naming neither prep supports nothing
    assert not entails("total RNA was extracted", "library.prep_type", "single-nucleus")


def test_prep_type_normalized_entailment_is_scoped_and_does_not_leak_to_quantification() -> None:
    """The normalized-prep shortcut is `library.prep_type` ONLY. The same biological prose must still
    never entail a counting decision — the firewall the parse/count split depends on.
    """
    from seqforge.harvest.verify import entails

    assert entails("single nuclei were profiled", "library.prep_type", "single-nucleus")
    assert not entails("single nuclei were profiled", "processing.quantification", "GeneFull")


def test_surface_forms_dispatch_is_exact_match_not_a_substring_test() -> None:
    """The old test was `if "chemistry" in field or "assay" in field` — it would misfire on any field
    that merely CONTAINS the word, e.g. `processing.assay_override`."""
    from seqforge.harvest.verify import surface_forms

    assert len(surface_forms("library.chemistry", "10x-3p-gex-v3")) > 1  # KB aliases attach
    assert surface_forms("processing.assay_override", "10x-3p-gex-v3") == ["10x-3p-gex-v3"]
    assert surface_forms("experiment.samples.tissue", "neurons") == ["neurons"]
    # quantification has NO alias source, on purpose: STARsolo's own spelling is the only form, so an
    # alias table could only loosen the one check that is not vacuous.
    assert surface_forms("processing.quantification", "GeneFull") == ["GeneFull"]


# ---------- PDF: pluggable extraction, unicode hardening, page tags, tables ----------
def test_the_default_pdf_backend_is_pymupdf() -> None:
    """The default was chosen by an end-to-end harvest eval (pymupdf read PDFs pypdf could not), not
    by license or speed. If this flips, the choice and its rationale must flip with it."""
    assert DEFAULT_PDF_BACKEND == "pymupdf"


@pytest.mark.parametrize("backend", ["pypdf", "pymupdf"])
def test_a_pdf_extracts_normalizes_and_verifies_end_to_end(
    tmp_path: Path, backend: PdfBackend
) -> None:
    """Both backends turn a PDF into canonical text a truthful quote greps back into, and the chosen
    backend is recorded on the document. The span space is the same contract as for a .txt."""
    pdf = _pdf(
        tmp_path,
        ["Methods. Libraries were prepared with the Chromium Single Cell 3' v3 kit."],
    )
    nd = normalize_document(pdf, pdf_backend=backend)
    assert nd.extractor == backend
    assert nd.pages and nd.pages[0].number == 1
    draft = AssertionDraft(
        field="library.chemistry",
        value="10x-3p-gex-v3",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="Chromium Single Cell 3' v3"),
        llm_confidence=0.9,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 1, report.rejected


def test_a_span_carries_the_physical_page_it_was_found_on(tmp_path: Path) -> None:
    """A PDF span is tagged with its 1-indexed page, computed by code from the offset — never by the
    model — so a citation can say "p.2" only where that is a real, checkable location."""
    pdf = _pdf(
        tmp_path,
        [
            "Methods. The data were deposited as PRJNA1027859 on submission.",
            "Results. The organism profiled was Caenorhabditis elegans throughout.",
        ],
    )
    nd = normalize_document(pdf, pdf_backend="pypdf")
    assert len(nd.pages) == 2
    drafts = [
        AssertionDraft(
            field="experiment.accessions",
            value="PRJNA1027859",
            span=SourceSpan(doc_sha256=nd.doc_sha256, quote="deposited as PRJNA1027859"),
            llm_confidence=0.9,
        ),
        AssertionDraft(
            field="experiment.organism",
            value="Caenorhabditis elegans",
            span=SourceSpan(doc_sha256=nd.doc_sha256, quote="Caenorhabditis elegans"),
            llm_confidence=0.9,
        ),
    ]
    report = verify_drafts(drafts, [nd], extractor=EXTRACTOR)
    by_field = {a.field: a for a in report.assertions}
    assert by_field["experiment.accessions"].span.page == 1  # page 1's accession
    assert by_field["experiment.organism"].span.page == 2  # page 2's organism


def test_a_non_pdf_span_has_no_page(tmp_path: Path) -> None:
    """ "p.4" is meaningless for a .txt or a record, so an unpaged source tags the span ``None`` rather
    than inventing a page."""
    nd = _doc(tmp_path, "The organism was Caenorhabditis elegans.")
    draft = AssertionDraft(
        field="experiment.organism",
        value="Caenorhabditis elegans",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="Caenorhabditis elegans"),
        llm_confidence=0.9,
    )
    a = verify_drafts([draft], [nd], extractor=EXTRACTOR).assertions[0]
    assert a.span.page is None and nd.pages == ()


def test_normalize_scrubs_invalid_unicode_that_would_crash_the_hash() -> None:
    """A PDF extractor emits NUL and orphaned UTF-16 surrogates on a bad font; a lone surrogate makes
    the very next step — ``text.encode()`` for the content hash — raise. The scrub runs first."""
    dirty = "Chromium Single\x00 Cell 3'\udc3c v3"  # NUL + lone low surrogate
    assert clean_invalid_unicode(dirty) == "Chromium Single Cell 3' v3"
    out = normalize_text(dirty)
    assert "Chromium Single Cell 3' v3" in out
    out.encode()  # would raise UnicodeEncodeError on a surviving surrogate; it does not


def test_normalize_document_survives_a_nul_byte_in_a_source_file(tmp_path: Path) -> None:
    """The end-to-end version: a file whose bytes contain a NUL still normalizes and hashes."""
    p = tmp_path / "methods.txt"
    p.write_bytes(b"We used the Chromium Single Cell 3' v3 kit.\x00\n")
    nd = normalize_document(p)
    assert "Chromium Single Cell 3' v3" in nd.text
    assert len(nd.normalized_sha256) == 64


def test_an_empty_pdf_is_refused_not_silently_empty(tmp_path: Path) -> None:
    """A scanned or image-only PDF yields no text. Refusing at the boundary is the difference between
    a clear cause and a mysterious span-verification miss that gets blamed on the model."""
    pdf = _pdf(tmp_path, [""])
    with pytest.raises(UnreadableDocument):
        normalize_document(pdf)


def test_garble_is_refused_but_a_short_instruction_is_not(tmp_path: Path) -> None:
    """The entropy gate catches a garbled read only past a length floor, so a terse legitimate
    instruction is never mistaken for noise."""
    garbled = tmp_path / "garbled.txt"
    garbled.write_text("x" * 600)  # long + near-zero entropy = a failed decode
    with pytest.raises(UnreadableDocument):
        normalize_document(garbled)

    terse = tmp_path / "note.txt"
    terse.write_text("use GeneFull")  # short + low-variety, but a real instruction
    assert normalize_document(terse, role="instruction").text == "use GeneFull"


def test_a_pdf_table_is_spliced_into_the_canonical_text_and_a_cell_verifies(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A detected table is rendered to ` | `-joined markdown and spliced INTO the canonical text (not a
    side channel), so a quoted sample-table cell greps back like any other span. The detection is
    pdfplumber's (exercised on real papers by the eval); here we pin OUR splice+render+verify path by
    supplying the table, so the test does not depend on pdfplumber's heuristics on a synthetic PDF."""
    monkeypatch.setattr(
        "seqforge.harvest.normalize._pdf_tables",
        lambda _p: {1: ["Sample | Genotype\n\nSRR1 | daf-2(e1370)"]},
    )
    pdf = _pdf(tmp_path, ["Methods. Per-sample genotypes are given in Table 1."])
    nd = normalize_document(pdf, pdf_backend="pypdf")
    assert "Table (page 1):" in nd.text
    assert "SRR1 | daf-2(e1370)" in nd.text  # the row survived normalize_text intact
    draft = AssertionDraft(
        field="experiment.samples.genotype",
        value="daf-2(e1370)",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="SRR1 | daf-2(e1370)"),
        llm_confidence=0.9,
    )
    assert verify_drafts([draft], [nd], extractor=EXTRACTOR).n_accepted == 1


def test_verify_rejects_a_chemistry_draft_that_names_no_kb_node(tmp_path: Path) -> None:
    """A chemistry value must NAME a chemistry. Span verification alone cannot ask that.

    `entails` is powerful exactly while the value comes from a controlled vocabulary — and nothing
    made it come from one. A model handed an SRA experiment record answers `library.chemistry` with
    the record's own `library_strategy`, "RNA-Seq", quoting it verbatim; the quote greps back, and
    entailment is *vacuous* because the value is a substring of its own quote. The draft passes both
    tripwires while claiming nothing, and then steers the resolver: asserted-bulk against an observed
    single-cell library is a cross-family conflict, so one such draft turned a decided dataset into a
    question and displaced a real hypothesis on another (#184).

    So code checks the half nothing else does — that the value resolves to a KB node — and that check
    is the vocabulary the entailment check was assuming all along.
    """
    nd = _doc(tmp_path, "GSM7486857: acr-2::GFP wt MNs rep 1; Caenorhabditis elegans; RNA-Seq")
    draft = AssertionDraft(
        field="library.chemistry",
        value="RNA-Seq",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="RNA-Seq"),
        llm_confidence=0.95,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 0
    assert report.rejected[0]["reason"] == "chemistry_names_no_kb_node"

    # ...and prove it is the VOCABULARY talking: both span-verification checks pass on this draft.
    from seqforge.harvest.verify import entails, find_span

    assert find_span(nd.text, draft.span.quote) is not None
    assert entails(draft.span.quote, draft.field, draft.value)


def test_verify_keeps_a_chemistry_draft_whose_value_names_a_node(tmp_path: Path) -> None:
    """The other half: the check must not cost the claims the channel exists for.

    A KB id and a real kit name both name a node, so both survive — the rejection is aimed at the
    term that names a whole field of assays, not at prose.
    """
    nd = _doc(
        tmp_path,
        "Libraries were prepared with the Chromium Next GEM Single-Cell 5' Reagent Kit v2 "
        "and the Chromium Single Cell 3' v3 kit.",
    )
    ids = AssertionDraft(
        field="library.chemistry",
        value="10x-3p-gex-v3",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="Chromium Single Cell 3' v3"),
        llm_confidence=0.9,
    )
    prose = AssertionDraft(
        field="library.chemistry",
        value="Chromium Next GEM Single-Cell 5' Reagent Kit v2",
        span=SourceSpan(
            doc_sha256=nd.doc_sha256, quote="Chromium Next GEM Single-Cell 5' Reagent Kit v2"
        ),
        llm_confidence=0.9,
    )
    report = verify_drafts([ids, prose], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 2, report.rejected


def test_the_kb_node_check_is_only_asked_of_the_chemistry_field(tmp_path: Path) -> None:
    """Every other field's vocabulary is its own; a strain is not a chemistry and owes the KB nothing."""
    nd = _doc(tmp_path, "Worms of strain RNA-Seq-1 were grown at 20C.")
    draft = AssertionDraft(
        field="experiment.samples.strain",
        value="RNA-Seq-1",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="strain RNA-Seq-1"),
        llm_confidence=0.9,
    )
    report = verify_drafts([draft], [nd], extractor=EXTRACTOR)
    assert report.n_accepted == 1, report.rejected


# ---------- mark-with-proof: a record re-rendering its own typed column is not a second source ----
#
# An archive record carries a fact twice — once as a typed column, and once inside the free text it
# renders for a human. SRA's experiment title is the shape that measured it: `library_strategy` is
# typed as `RNA-Seq`, and the same three characters end the title the model is shown. A model asked
# for the chemistry answers with them, quoting verbatim; the quote greps back and entailment is
# vacuous, so both tripwires pass on a claim that is a transcription of a column code already read.
#
# So the span is marked where it can be PROVED — byte-equal to a value typed on that same record —
# and a draft whose quote lies wholly inside a marked span is refused. It is the same principle as
# `resolve/records.py`'s declared-beats-read, one layer earlier: one deposit is one source, at every
# layer (ADR-0021). Archive-neutral by construction: no delimiter grammar, no accession format, and
# nothing at all to do where there are no records.


def _experiment_record(
    *, title: str = "GSM1: MC38 tumor 3 weeks; Mus musculus; RNA-Seq", strategy: str = "RNA-Seq"
) -> ArchiveRecord:
    """One SRA experiment record: a typed column, and the title that renders it back out."""
    return ArchiveRecord(
        level="experiment",
        accession="SRX1",
        parent="SAMN1",
        attributes=[RecordAttribute(name="library_strategy", value=strategy)],
        free_text=[FreeText(label="experiment_title", text=title)],
    )


def test_a_span_that_re_renders_the_records_own_typed_column_is_marked() -> None:
    """The mark carries its own proof: which attribute, and the value that matched."""
    from seqforge.harvest import normalize_record

    doc = normalize_record(_experiment_record())
    assert [(doc.text[d.start : d.end], d.attribute, d.value) for d in doc.declared] == [
        ("RNA-Seq", "library_strategy", "RNA-Seq")
    ]


def test_a_typed_value_appearing_only_inside_a_longer_token_is_not_marked() -> None:
    """Whole words, never characters — the `CD4` inside `CD45` caution, one layer earlier."""
    from seqforge.harvest import normalize_record

    doc = normalize_record(
        _experiment_record(title="GSM1: CD45 positive cells; Mus musculus", strategy="CD4")
    )
    assert doc.declared == ()


def test_verify_rejects_a_draft_quoting_the_records_own_typed_column() -> None:
    """The measured shape: `library.chemistry = "RNA-Seq"`, quoted out of the title that renders it.

    Both existing tripwires pass on this draft — the quote is real and the value sits inside it — and
    it is refused anyway, because the span is the record repeating a column, not prose about it.
    """
    from seqforge.harvest import normalize_record

    doc = normalize_record(_experiment_record())
    draft = AssertionDraft(
        field="library.chemistry",
        value="RNA-Seq",
        span=SourceSpan(doc_sha256=doc.doc_sha256, quote="RNA-Seq"),
        llm_confidence=0.95,
    )
    report = verify_drafts([draft], [doc], extractor=EXTRACTOR)
    assert report.n_accepted == 0
    assert report.rejected[0]["reason"] == "quote_is_a_typed_column"
    assert "library_strategy" in str(report.rejected[0]["detail"])

    # ...and it is the MARK talking, not the two checks that already existed.
    assert find_span(doc.text, draft.span.quote) is not None
    assert entails(draft.span.quote, draft.field, draft.value)


def test_a_quote_reaching_past_the_typed_column_is_still_prose() -> None:
    """ "Wholly inside" is the whole rule: a quote that says more than the column did is a reading.

    Both drafts here would die if the mark rejected anything merely *touching* it — one quotes a
    different part of the same title, the other quotes the title entire.
    """
    from seqforge.harvest import normalize_record

    doc = normalize_record(_experiment_record())
    part = AssertionDraft(
        field="experiment.samples.treatment",
        value="MC38 tumor 3 weeks",
        span=SourceSpan(doc_sha256=doc.doc_sha256, quote="MC38 tumor 3 weeks"),
        llm_confidence=0.9,
    )
    whole = AssertionDraft(
        field="experiment.samples.treatment",
        value="MC38 tumor 3 weeks",
        span=SourceSpan(doc_sha256=doc.doc_sha256, quote=doc.text.splitlines()[-1]),
        llm_confidence=0.9,
    )
    report = verify_drafts([part, whole], [doc], extractor=EXTRACTOR)
    assert report.n_accepted == 2, report.rejected


def test_a_document_that_came_from_no_record_carries_no_marks(tmp_path: Path) -> None:
    """The no-op that keeps this archive-neutral: a paper someone handed us has no typed columns.

    Most sequencing data never had an accession, so a rule that needed one would be a rule most
    datasets could not obey. There is nothing to mark here and nothing is marked.
    """
    nd = _doc(tmp_path, "We prepared libraries with the Chromium Single Cell 3' v3 kit. RNA-Seq.")
    assert nd.declared == ()
    draft = AssertionDraft(
        field="library.chemistry",
        value="10x-3p-gex-v3",
        span=SourceSpan(doc_sha256=nd.doc_sha256, quote="Chromium Single Cell 3' v3"),
        llm_confidence=0.9,
    )
    assert verify_drafts([draft], [nd], extractor=EXTRACTOR).n_accepted == 1


def test_a_collapsed_run_document_marks_the_columns_its_members_typed() -> None:
    """The runs of one sample become one document, and the marks have to survive the concatenation.

    Offsets are computed against the joined canonical text, so a mark found in the second run's
    rendering must point into the joined string rather than into that run's own.
    """
    from seqforge.harvest import plan_extraction
    from seqforge.models.records import ArchiveRecordSet

    runs = [
        ArchiveRecord(
            level="run",
            accession=f"SRR{i}",
            parent="SAMN1",
            attributes=[RecordAttribute(name="library_strategy", value="RNA-Seq")],
            free_text=[FreeText(label="run_alias", text=f"MC38 tumor 3 weeks rep{i}; RNA-Seq")],
        )
        for i in (1, 2)
    ]
    records = ArchiveRecordSet(
        source="fixture",
        query="PRJNA1",
        records=[ArchiveRecord(level="sample", accession="SAMN1"), *runs],
    )
    doc = next(d for d in plan_extraction(records=records).documents if d.scope == "run")
    assert [doc.text[d.start : d.end] for d in doc.declared] == ["RNA-Seq", "RNA-Seq"]
    assert len({d.start for d in doc.declared}) == 2, "each occurrence has its own offset"


# ---------- the invariant: what a group of near-identical records shares, at TOKEN granularity ----
# A second kind of mark, and it must never become the first. `DeclaredSpan` makes `verify` REFUSE a
# quote; a `VariantSpan` decides whether a claim FANS. Conflate them and a fanned claim is silently
# rejected, or a quote inside a column is fanned into records that never carried it.


def test_the_invariant_is_computed_at_token_boundaries_and_age_3_never_fans_into_age_30() -> None:
    """The hazard the whole granularity decision exists for, pinned.

    At CHARACTER granularity ``3`` is a shared span between ``age: 3`` and ``age: 30``, so ``age = 3``
    would fan into the record that says 30 and the leftover variant would be the single character
    ``0`` — which no askable-content test would ever send. A wrong claim, silently, on the one record
    that differed. At token granularity ``3`` and ``30`` are two whole tokens, the position varies,
    and nothing about it is shared.
    """
    from seqforge.harvest import is_invariant_span, variant_spans, varying_token_indices

    texts = ["sample S1\n\nage: 3", "sample S2\n\nage: 3", "sample S3\n\nage: 30"]

    assert varying_token_indices(texts) == (1, 3), "the accession token and the age token"
    marks = variant_spans(texts)
    assert [(m.value, m.n_values) for m in marks] == [("S1", 3), ("3", 2)]
    assert texts[0][marks[1].start : marks[1].end] == "3", (
        "the WHOLE token, never a character slice"
    )

    doc = NormalizedDoc(
        doc_sha256="d", normalized_sha256="d", text=texts[0], source_basename="s.txt",
        variants=marks,
    )  # fmt: skip
    start = texts[0].index("age: 3")
    assert not is_invariant_span(doc, start, start + len("age: 3")), "reaches into the variant"
    assert is_invariant_span(doc, start, start + len("age:")), "the label alone is shared"


def test_a_serial_name_holding_an_underscore_is_one_token_and_therefore_varies() -> None:
    """`_is_token_char` counts `_`, so `nasal_prox1_270` is ONE token — the measured consequence being
    that a serially-named deposit's sample documents do not collapse and every one of them is asked.

    That is the design pricing itself, not a defect to fix. The rescue that was tried and rejected —
    mask a digit run whose values across the group are exactly an enumeration `1..N` — works on a
    plate and dies on `sample_1h ... sample_24h`, where the digit IS the data.
    """
    from seqforge.harvest import variant_spans

    marks = variant_spans(["sample_title: nasal_prox1_270", "sample_title: nasal_prox1_271"])

    assert [m.value for m in marks] == ["nasal_prox1_270"], "one token, not a shared 12-char prefix"
    assert marks[0].end - marks[0].start == len("nasal_prox1_270")


def test_a_digit_that_is_the_data_varies_exactly_as_a_word_does() -> None:
    """`sample_1h` vs `sample_24h` is the case that killed the enumeration rescue, and the rule that
    replaced it needs no test to tell it from a well index: both are whole tokens, both vary."""
    from seqforge.harvest import varying_token_indices

    plate = ["well A1", "well A2", "well A3"]
    timecourse = ["dose sample_1h", "dose sample_6h", "dose sample_24h"]

    assert varying_token_indices(plate) == (1,)
    assert varying_token_indices(timecourse) == (1,), (
        "no code-side test separates these, and none is"
    )


def test_texts_of_different_shapes_are_not_one_group_and_refuse_to_be_aligned() -> None:
    """A skeleton is a whole-string equality key, deliberately: an edit-distance alignment would make
    "near-identical" a threshold, and a threshold is a knob nobody can tune from bytes. A record whose
    paragraph is shaped differently forms its own group and is asked on its own."""
    from seqforge.harvest import token_skeleton, varying_token_indices

    assert token_skeleton("age: 3") != token_skeleton("age = 3"), "punctuation is part of the shape"
    assert token_skeleton("age: 3") != token_skeleton("age: 3 days"), "so is the token count"

    with pytest.raises(ValueError, match="one skeleton"):
        varying_token_indices(["age: 3", "age: 3 days"])


def test_an_unmarked_document_carries_no_group_and_no_claim_of_one() -> None:
    """Every document a human handed us, and every record read on its own. `is_invariant_span` is
    vacuously true there, and the CALLER — never this predicate — is what knows there is nothing to
    fan to."""
    from seqforge.harvest import is_invariant_span, variant_spans

    assert variant_spans(["one text only"]) == ()
    doc = NormalizedDoc(doc_sha256="d", normalized_sha256="d", text="abc", source_basename="s.txt")
    assert doc.variants == () and is_invariant_span(doc, 0, 3)


def test_a_variant_document_keeps_the_accession_and_drops_the_labels() -> None:
    """What a non-exemplar member is sent as: its distinctive bytes, and nothing it shares.

    The accession leads because it *varies* and everything that varies is kept — no special case. It
    earns that place twice over: a document's identity is its bytes, so two members that happen to
    share a serial name would otherwise reduce to one `doc_sha256`, and `resolve` keys a document's
    subject by that sha — one sample would silently inherit the other's claims.
    """
    from seqforge.harvest import variant_text, varying_token_indices

    texts = [
        f"sample SAMN{i}\n\nsample_title: nasal_prox1_{i}\n\nsample_alias: GSM627690{i}"
        for i in (1, 2)
    ]
    varying = varying_token_indices(texts)

    assert variant_text(texts[0], varying) == "SAMN1\n\nnasal_prox1_1\n\nGSM6276901"
    assert variant_text(texts[0], varying) != variant_text(texts[1], varying)


def test_adjacent_variants_keep_the_separator_the_record_wrote_between_them() -> None:
    """Adjacency is the only syntax the group did NOT already read in the exemplar, so it is the only
    syntax worth carrying — and a paragraph break between tokens that were apart is what stops the
    join from inventing a phrase no record wrote (`genotype: WT` + `age: 72` must not read `WT 72`).
    """
    from seqforge.harvest import variant_text, varying_token_indices

    titles = [f"experiment SRX{i}\n\ntitle: GSM{i}: nasal_prox1_{i}; Mus musculus" for i in (1, 2)]
    genotypes = [f"sample SAMN{i}\n\ngenotype: {g}\n\nage: {a}" for i, (g, a) in
                 enumerate([("WT", "72"), ("mut", "24")], start=1)]  # fmt: skip

    assert variant_text(titles[0], varying_token_indices(titles)) == "SRX1\n\nGSM1: nasal_prox1_1"
    assert variant_text(genotypes[0], varying_token_indices(genotypes)) == "SAMN1\n\nWT\n\n72"


def test_a_variant_document_is_regenerable_from_the_record_set_and_only_from_it() -> None:
    """ADR-0030's subject, made concrete. `render_record` promises a document is reproducible from one
    record; this text is not — it needs the record set, because which tokens vary is a property of the
    group. Same records in, same bytes out, forever; one record alone cannot produce it."""
    from seqforge.harvest import variant_text, varying_token_indices

    texts = [f"run SRR{i}\n\nrun_alias: N2_wild_type_r{i}" for i in (1, 2, 3)]
    once = variant_text(texts[0], varying_token_indices(texts))

    assert once == variant_text(texts[0], varying_token_indices(list(texts))), "deterministic"
    assert varying_token_indices(texts[:1]) == (), "one record alone says nothing varies at all"
    # ...and a DIFFERENT set gives a different reduction of the same record, which is exactly why the
    # record set and not the record is the unit of reproducibility here.
    assert once == "SRR1\n\nN2_wild_type_r1", "the whole alias is one token, so all of it varies"
    strains = ["run SRR1\n\nrun_alias: N2 rep 1", "run SRR2\n\nrun_alias: daf2 rep 1"]
    indices = ["run SRR1\n\nrun_alias: N2 rep 1", "run SRR2\n\nrun_alias: N2 rep 2"]
    assert variant_text(strains[0], varying_token_indices(strains)) == "SRR1\n\nN2"
    assert variant_text(indices[0], varying_token_indices(indices)) == "SRR1\n\n1"
